"""Сервис экспорта данных в Excel (ФТ-09-04)."""

from __future__ import annotations

from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def render_xlsx(
    *,
    sheet_title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence],
    column_widths: Sequence[int] | None = None,
) -> bytes:
    """Универсальный рендер таблицы в xlsx с заголовком и стилизацией."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_title[:31]  # лимит листа в Excel

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2563EB", end_color="2563EB", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Заголовки
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Данные
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Ширина колонок
    if column_widths:
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    else:
        # Авто-ширина (приблизительно)
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
